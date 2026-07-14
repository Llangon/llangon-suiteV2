from __future__ import annotations

import copy
import hashlib
import io
import sqlite3
import zipfile
from datetime import datetime, timezone
from pathlib import Path

import pytest
from openpyxl import load_workbook
from PIL import Image

from webapp.infonalia_webapp.justificaciones_baja.application.errors import (
    JustificationConflictApplicationError,
    JustificationValidationError,
)
from webapp.infonalia_webapp.justificaciones_baja.application.service import (
    JustificationApplicationService,
)
from webapp.infonalia_webapp.justificaciones_baja.persistence import (
    JustificationConflictError,
    JustificationRepository,
    ensure_justificaciones_baja_schema,
)


FIXED_TIME = datetime(2026, 7, 14, 10, 30, tzinfo=timezone.utc)


class LowestRandom:
    def randint(self, minimum: int, maximum: int) -> int:
        return minimum


def _database() -> sqlite3.Connection:
    conn = sqlite3.connect(":memory:")
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA foreign_keys = ON")
    conn.execute(
        """
        CREATE TABLE licitaciones (
            id INTEGER PRIMARY KEY,
            expediente TEXT NOT NULL,
            objeto TEXT NOT NULL,
            organismo TEXT NOT NULL,
            ruta_carpeta TEXT NOT NULL DEFAULT ''
        )
        """
    )
    conn.execute(
        """
        CREATE TABLE clientes (
            id INTEGER PRIMARY KEY,
            razon_social TEXT NOT NULL
        )
        """
    )
    conn.execute(
        "INSERT INTO licitaciones VALUES (1, 'EXP-TEST', 'Suministro de prueba', 'Organismo de prueba', '2026/EXP-TEST')"
    )
    conn.execute("INSERT INTO clientes VALUES (1, 'Empresa de prueba, S.L.')")
    ensure_justificaciones_baja_schema(conn)
    return conn


def _service(conn: sqlite3.Connection, tmp_path: Path) -> JustificationApplicationService:
    return JustificationApplicationService(
        JustificationRepository(conn),
        clock=lambda: FIXED_TIME,
        random_source=LowestRandom(),
        temporary_root=tmp_path / "runtime",
    )


def _create(service: JustificationApplicationService) -> dict:
    return service.create(
        licitacion={
            "id": 1,
            "expediente": "EXP-TEST",
            "objeto": "Suministro de prueba",
            "organismo": "Organismo de prueba",
        },
        cliente={
            "id": 1,
            "razon_social": "Empresa de prueba, S.L.",
            "nif_cif": "B00000000",
            "domicilio_fiscal": "Calle de prueba 1",
            "municipio": "Sevilla",
            "representante_nombre": "Representante de prueba",
        },
        lote_numero="1",
        lote_nombre="Alimentos",
        declared_offer="1000",
        user_id="admin",
    )


def _complete_draft(item: dict) -> dict:
    draft = item["draft"]
    draft["transport"].update(
        {
            "operational_weeks": 10,
            "weekly_deliveries": 2,
            "circular_kilometres": "50",
            "effective_decimal_hours": "2.5",
            "kilometre_rate": "0.6",
            "hourly_rate": "30",
            "contract_stops": 1,
            "shared_orders": 10,
            "route_duration_text": "2 h 30 min",
        }
    )
    draft["financial"].update(
        {
            "declared_lot_offer": "1000",
            "general_expense_base": "1000",
            "general_expense_percentage": "0.05",
            "indirect_costs": "10",
        }
    )
    draft["transport_document"].update(
        {
            "observatory_date": "Abril de 2026",
            "observatory_url": "https://example.invalid/observatorio.pdf",
        }
    )
    draft["narrative"].update(
        {
            "exposition": "Se presenta una estimación económica para validación.",
            "arguments": ["Organización logística consolidada."],
            "acquisition_text": "Los costes de adquisición son estimativos.",
            "transport_text": "La ruta se integra en la distribución habitual.",
            "structure_text": "La estructura existente absorbe el contrato.",
            "conclusion": "La oferta resulta económicamente viable.",
        }
    )
    draft["products"] = [
        {
            "line_id": "line-a",
            "name": "Producto repetido",
            "characteristics": "Formato A",
            "quantity": "5",
            "offered_unit_price": "100",
            "locked": False,
            "cost_origin": "sin_generar",
        },
        {
            "line_id": "line-b",
            "name": "Producto repetido",
            "characteristics": "Formato A",
            "quantity": "10",
            "offered_unit_price": "50",
            "locked": False,
            "cost_origin": "sin_generar",
        },
    ]
    return draft


def test_costs_are_only_generated_explicitly_and_survive_reopen(tmp_path: Path) -> None:
    conn = _database()
    service = _service(conn, tmp_path)
    created = _create(service)
    saved = service.save(
        created["id"],
        draft=_complete_draft(created),
        expected_revision=created["revision"],
        user_id="admin",
    )
    assert all(item["generated_unit_cost"] is None for item in saved["draft"]["products"])

    generated = service.generate_costs(
        created["id"], expected_revision=saved["revision"], user_id="admin"
    )
    frozen_costs = [item["generated_unit_cost"] for item in generated["draft"]["products"]]
    assert frozen_costs == ["71.429", "35.714"]

    reopened = service.get(created["id"])
    assert [item["generated_unit_cost"] for item in reopened["draft"]["products"]] == frozen_costs
    assert reopened["calculation"]["values"]["raw"]["declared_lot_offer"] == "1000"


def test_economic_errors_are_structured_and_failed_freeze_is_read_only(
    tmp_path: Path,
) -> None:
    conn = _database()
    service = _service(conn, tmp_path)
    created = _create(service)
    draft = _complete_draft(created)

    preview = service.preview(draft)
    preview_errors = preview["calculation"]["errors"]
    assert {item["code"] for item in preview_errors} == {"producto_sin_coste_efectivo"}
    assert all(
        set(item) == {"code", "severity", "message", "field", "line_id", "metadata"}
        for item in preview_errors
    )
    assert {item["severity"] for item in preview_errors} == {"error"}
    assert {item["line_id"] for item in preview_errors} == {"line-a", "line-b"}

    saved = service.save(
        created["id"],
        draft=draft,
        expected_revision=created["revision"],
        user_id="admin",
    )
    assert saved["calculation"]["errors"] == preview_errors
    before_revision = saved["revision"]
    before_history = list(saved["history"])

    with pytest.raises(JustificationValidationError) as raised:
        service.freeze(
            created["id"],
            expected_revision=before_revision,
            user_id="admin",
        )

    assert raised.value.code == "validacion_justificacion"
    assert raised.value.issues == preview_errors
    reopened = service.get(created["id"])
    assert reopened["revision"] == before_revision
    assert reopened["draft_frozen"] is False
    assert reopened["versions"] == []
    assert reopened["history"] == before_history


def test_client_change_updates_draft_record_filters_and_history_atomically(
    tmp_path: Path,
) -> None:
    conn = _database()
    conn.execute("INSERT INTO clientes VALUES (2, 'Segundo cliente de prueba, S.L.')")
    service = _service(conn, tmp_path)
    created = _create(service)
    draft = _complete_draft(created)
    draft["client"].update(
        {
            "client_id": 2,
            "name": "Segundo cliente de prueba, S.L.",
            "nif": "B00000001",
            "address": "Calle de prueba 2",
        }
    )

    changed = service.save(
        created["id"],
        draft=draft,
        expected_revision=created["revision"],
        user_id="admin",
    )

    assert changed["cliente_id"] == 2
    assert changed["cliente_razon_social"] == "Segundo cliente de prueba, S.L."
    assert changed["draft"]["client"]["client_id"] == 2
    assert [item["id"] for item in service.list(cliente_id=2)] == [created["id"]]
    assert service.list(cliente_id=1) == []
    client_events = [
        item for item in changed["history"] if item["event_type"] == "client_changed"
    ]
    assert len(client_events) == 1
    assert client_events[0]["metadata"] == {
        "cliente_id": 2,
        "previous_cliente_id": 1,
    }

    invalid = copy.deepcopy(changed["draft"])
    invalid["client"]["client_id"] = 999
    with pytest.raises(JustificationValidationError, match="cliente"):
        service.save(
            created["id"],
            draft=invalid,
            expected_revision=changed["revision"],
            user_id="admin",
        )

    reopened = service.get(created["id"])
    assert reopened["revision"] == changed["revision"]
    assert reopened["cliente_id"] == 2
    assert reopened["draft"] == changed["draft"]
    assert reopened["history"] == changed["history"]


def test_bulk_product_lock_is_one_atomic_revision_and_rolls_back_on_failure(
    tmp_path: Path,
) -> None:
    conn = _database()
    service = _service(conn, tmp_path)
    created = _create(service)
    saved = service.save(
        created["id"],
        draft=_complete_draft(created),
        expected_revision=created["revision"],
        user_id="admin",
    )
    before = service.get(created["id"])

    with pytest.raises(JustificationValidationError):
        service.set_product_locks(
            created["id"],
            expected_revision=before["revision"],
            line_ids=["line-a", "line-inexistente"],
            locked=True,
            user_id="admin",
        )
    after_domain_failure = service.get(created["id"])
    assert after_domain_failure["revision"] == before["revision"]
    assert after_domain_failure["draft"] == before["draft"]
    assert after_domain_failure["history"] == before["history"]

    conn.execute(
        """
        CREATE TRIGGER reject_products_locked
        BEFORE INSERT ON justificacion_baja_historial
        WHEN NEW.event_type = 'products_locked'
        BEGIN
            SELECT RAISE(ABORT, 'forced bulk lock history failure');
        END
        """
    )
    with pytest.raises(sqlite3.IntegrityError, match="forced bulk lock history failure"):
        service.set_product_locks(
            created["id"],
            expected_revision=before["revision"],
            line_ids=["line-a", "line-b"],
            locked=True,
            user_id="admin",
        )
    conn.execute("DROP TRIGGER reject_products_locked")
    after_storage_failure = service.get(created["id"])
    assert after_storage_failure["revision"] == before["revision"]
    assert after_storage_failure["draft"] == before["draft"]
    assert after_storage_failure["history"] == before["history"]

    locked = service.set_product_locks(
        created["id"],
        expected_revision=before["revision"],
        line_ids=["line-a", "line-b", "line-a"],
        locked=True,
        user_id="admin",
    )
    assert locked["revision"] == before["revision"] + 1
    assert all(item["locked"] is True for item in locked["draft"]["products"])
    lock_events = [
        item for item in locked["history"] if item["event_type"] == "products_locked"
    ]
    assert len(lock_events) == 1
    assert lock_events[0]["metadata"]["line_ids"] == ["line-a", "line-b"]
    assert lock_events[0]["metadata"]["locked"] is True


def test_manual_lock_and_selective_recalculation_are_traced(tmp_path: Path) -> None:
    conn = _database()
    service = _service(conn, tmp_path)
    created = _create(service)
    saved = service.save(
        created["id"], draft=_complete_draft(created), expected_revision=1, user_id="admin"
    )
    generated = service.generate_costs(
        created["id"], expected_revision=saved["revision"], user_id="admin"
    )
    manual = service.set_manual_cost(
        created["id"],
        expected_revision=generated["revision"],
        line_id="line-a",
        manual_unit_cost="60,500",
        user_id="admin",
    )
    locked = service.set_product_lock(
        created["id"],
        expected_revision=manual["revision"],
        line_id="line-b",
        locked=True,
        user_id="admin",
    )
    before = {item["line_id"]: dict(item) for item in locked["draft"]["products"]}
    recalculated = service.recalculate_costs(
        created["id"],
        expected_revision=locked["revision"],
        line_ids=None,
        user_id="admin",
    )
    after = {item["line_id"]: item for item in recalculated["draft"]["products"]}
    assert after["line-a"]["manual_unit_cost"] == "60.500"
    assert after["line-a"]["cost_origin"] == "manual"
    assert after["line-b"] == before["line-b"]
    event_types = {item["event_type"] for item in recalculated["history"]}
    assert {"costs_generated", "manual_cost_set", "product_locked", "costs_recalculated"} <= event_types


def test_freeze_and_generate_word_excel_from_same_snapshot(tmp_path: Path) -> None:
    conn = _database()
    service = _service(conn, tmp_path)
    created = _create(service)
    saved = service.save(
        created["id"], draft=_complete_draft(created), expected_revision=1, user_id="admin"
    )
    generated = service.generate_costs(
        created["id"], expected_revision=saved["revision"], user_id="admin"
    )
    frozen = service.freeze(
        created["id"], expected_revision=generated["revision"], user_id="admin"
    )
    base = tmp_path / "dropbox"
    output = base / "2026" / "EXP-TEST" / "Justificaciones de baja" / "Lote_1"
    base.mkdir(parents=True)
    result = service.generate_documents(
        created["id"],
        version_number=frozen["version"]["version_number"],
        output_directory=output,
        dropbox_base=base,
        user_id="admin",
    )
    assert len(result["documents"]) == 2
    assert len({item["payload_sha256"] for item in result["documents"]}) == 1
    assert len({item["version_id"] for item in result["documents"]}) == 1
    for document in result["documents"]:
        path = base / Path(document["relative_path"])
        assert path.exists()
        assert hashlib.sha256(path.read_bytes()).hexdigest().upper() == document["sha256"]
    workbook_path = next(base.rglob("*.xlsx"))
    workbook = load_workbook(workbook_path, data_only=False)
    try:
        formulas = [
            cell.value
            for sheet in workbook.worksheets
            for row in sheet.iter_rows()
            for cell in row
            if isinstance(cell.value, str) and cell.value.startswith("=")
        ]
    finally:
        workbook.close()
    assert formulas == []
    word_path = next(base.rglob("*.docx"))
    with zipfile.ZipFile(word_path) as archive:
        all_xml = b"".join(
            archive.read(name) for name in archive.namelist() if name.endswith(".xml")
        )
        story_xml = b"".join(
            archive.read(name)
            for name in archive.namelist()
            if name == "word/document.xml"
            or name.startswith("word/header")
            or name.startswith("word/footer")
        )
    assert b"{{" not in all_xml and b"{%" not in all_xml
    assert b"instrText" not in story_xml and b"fldSimple" not in story_xml


def test_changing_range_and_saving_never_regenerates_costs(tmp_path: Path) -> None:
    conn = _database()
    service = _service(conn, tmp_path)
    created = _create(service)
    saved = service.save(
        created["id"], draft=_complete_draft(created), expected_revision=1, user_id="admin"
    )
    generated = service.generate_costs(
        created["id"], expected_revision=saved["revision"], user_id="admin"
    )
    before = [item["generated_unit_cost"] for item in generated["draft"]["products"]]
    generated["draft"]["cost_range"] = {
        "minimum_percentage": 55,
        "maximum_percentage": 60,
    }
    changed = service.save(
        created["id"],
        draft=generated["draft"],
        expected_revision=generated["revision"],
        user_id="admin",
    )
    assert [item["generated_unit_cost"] for item in changed["draft"]["products"]] == before
    recalculated = service.recalculate_costs(
        created["id"],
        expected_revision=changed["revision"],
        line_ids=None,
        user_id="admin",
    )
    assert [item["generated_unit_cost"] for item in recalculated["draft"]["products"]] != before


def test_optimistic_revision_rejects_silent_overwrite(tmp_path: Path) -> None:
    conn = _database()
    service = _service(conn, tmp_path)
    created = _create(service)
    stale = created["revision"]
    service.save(
        created["id"], draft=_complete_draft(created), expected_revision=stale, user_id="admin"
    )
    try:
        service.save(
            created["id"], draft=_complete_draft(created), expected_revision=stale, user_id="admin"
        )
    except JustificationConflictApplicationError as exc:
        assert "otra pestaña" in str(exc)
    else:  # pragma: no cover - defensive assertion
        raise AssertionError("Se permitió sobrescribir con una revisión obsoleta.")


def test_route_image_is_hashed_and_new_version_keeps_frozen_snapshot(tmp_path: Path) -> None:
    conn = _database()
    service = _service(conn, tmp_path)
    created = _create(service)
    saved = service.save(
        created["id"], draft=_complete_draft(created), expected_revision=1, user_id="admin"
    )
    generated = service.generate_costs(
        created["id"], expected_revision=saved["revision"], user_id="admin"
    )
    image_buffer = io.BytesIO()
    Image.new("RGB", (320, 180), color=(230, 240, 250)).save(image_buffer, format="PNG")
    image_bytes = image_buffer.getvalue()
    attached = service.attach_route_image(
        created["id"],
        expected_revision=generated["revision"],
        filename="ruta-test.png",
        content=image_bytes,
        user_id="admin",
    )
    image_reference = attached["draft"]["route_image"]
    assert image_reference["sha256"] == hashlib.sha256(image_bytes).hexdigest().upper()
    assert "content" not in image_reference

    first = service.freeze(
        created["id"], expected_revision=attached["revision"], user_id="admin"
    )
    first_snapshot = first["version"]["snapshot_json"]
    edited = first["item"]["draft"]
    edited["narrative"]["conclusion"] = "Conclusión revisada para una segunda versión."
    next_draft = service.save(
        created["id"],
        draft=edited,
        expected_revision=first["item"]["revision"],
        user_id="admin",
    )
    assert next_draft["draft_frozen"] is False
    assert next_draft["draft_based_on_version"] == 1
    second = service.freeze(
        created["id"], expected_revision=next_draft["revision"], user_id="admin"
    )
    assert second["version"]["version_number"] == 2
    assert service.repository.get_version(created["id"], version_number=1)["snapshot_json"] == first_snapshot


def test_route_image_persistence_conflict_uses_application_409_contract(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    conn = _database()
    service = _service(conn, tmp_path)
    created = _create(service)
    image_buffer = io.BytesIO()
    Image.new("RGB", (64, 36), color=(230, 240, 250)).save(image_buffer, format="PNG")

    def fail_put_route_image(*args: object, **kwargs: object) -> dict:
        raise JustificationConflictError("Conflicto simulado al guardar la imagen.")

    monkeypatch.setattr(JustificationRepository, "put_route_image", fail_put_route_image)

    with pytest.raises(JustificationConflictApplicationError) as caught:
        service.attach_route_image(
            created["id"],
            expected_revision=created["revision"],
            filename="ruta.png",
            content=image_buffer.getvalue(),
            user_id="admin",
        )

    assert caught.value.status_code == 409
    assert caught.value.code == "conflicto_revision"
    assert "Conflicto simulado" in str(caught.value)
    runtime = tmp_path / "runtime"
    assert not runtime.exists() or list(runtime.iterdir()) == []


def test_route_image_rolls_back_asset_if_draft_update_conflicts(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    conn = _database()
    service = _service(conn, tmp_path)
    created = _create(service)
    image_buffer = io.BytesIO()
    Image.new("RGB", (64, 36), color=(230, 240, 250)).save(image_buffer, format="PNG")

    def fail_update_draft(*args: object, **kwargs: object) -> dict:
        raise JustificationConflictError("Conflicto posterior al alta del asset.")

    monkeypatch.setattr(JustificationRepository, "update_draft", fail_update_draft)

    with pytest.raises(JustificationConflictApplicationError):
        service.attach_route_image(
            created["id"],
            expected_revision=created["revision"],
            filename="ruta.png",
            content=image_buffer.getvalue(),
            user_id="admin",
        )

    owner = conn.execute(
        "SELECT route_asset_id, revision FROM justificaciones_baja WHERE id = ?",
        (created["id"],),
    ).fetchone()
    assert owner["route_asset_id"] is None
    assert owner["revision"] == created["revision"]
    assert conn.execute("SELECT COUNT(*) FROM justificacion_baja_assets").fetchone()[0] == 0
    assert conn.execute(
        "SELECT COUNT(*) FROM justificacion_baja_historial WHERE event_type LIKE 'route_image_%'"
    ).fetchone()[0] == 0


def test_document_registration_conflict_uses_application_409_contract_and_cleans_files(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    conn = _database()
    service = _service(conn, tmp_path)
    created = _create(service)
    saved = service.save(
        created["id"], draft=_complete_draft(created), expected_revision=1, user_id="admin"
    )
    generated = service.generate_costs(
        created["id"], expected_revision=saved["revision"], user_id="admin"
    )
    frozen = service.freeze(
        created["id"], expected_revision=generated["revision"], user_id="admin"
    )
    base = tmp_path / "dropbox"
    output = base / "2026" / "EXP-TEST" / "Justificaciones de baja" / "Lote_1"
    base.mkdir(parents=True)

    original_add_document = JustificationRepository.add_document
    calls = 0

    def fail_second_add_document(self: JustificationRepository, *args: object, **kwargs: object) -> dict:
        nonlocal calls
        calls += 1
        if calls == 2:
            raise JustificationConflictError("Conflicto simulado al registrar documentos.")
        return original_add_document(self, *args, **kwargs)

    monkeypatch.setattr(JustificationRepository, "add_document", fail_second_add_document)

    with pytest.raises(JustificationConflictApplicationError) as caught:
        service.generate_documents(
            created["id"],
            version_number=frozen["version"]["version_number"],
            output_directory=output,
            dropbox_base=base,
            user_id="admin",
        )

    assert caught.value.status_code == 409
    assert caught.value.code == "conflicto_revision"
    assert "Conflicto simulado" in str(caught.value)
    assert calls == 2
    assert list(base.rglob("*.docx")) == []
    assert list(base.rglob("*.xlsx")) == []
    assert conn.execute("SELECT COUNT(*) FROM justificacion_baja_documentos").fetchone()[0] == 0
    assert conn.execute(
        "SELECT COUNT(*) FROM justificacion_baja_historial WHERE event_type LIKE 'document_%'"
    ).fetchone()[0] == 0


def test_document_generation_skips_an_orphaned_filename_without_overwriting_it(
    tmp_path: Path,
) -> None:
    conn = _database()
    service = _service(conn, tmp_path)
    created = _create(service)
    saved = service.save(
        created["id"], draft=_complete_draft(created), expected_revision=1, user_id="admin"
    )
    generated = service.generate_costs(
        created["id"], expected_revision=saved["revision"], user_id="admin"
    )
    frozen = service.freeze(
        created["id"], expected_revision=generated["revision"], user_id="admin"
    )
    base = tmp_path / "dropbox"
    output = base / "EXP-TEST" / "Justificaciones de baja" / "Lote_1"
    version_directory = output / "Version_001"
    version_directory.mkdir(parents=True)
    orphan = version_directory / "Justificacion_Baja_EXP-TEST_Lote_1_v001.docx"
    orphan.write_bytes(b"orphan-must-not-be-overwritten")

    result = service.generate_documents(
        created["id"],
        version_number=frozen["version"]["version_number"],
        output_directory=output,
        dropbox_base=base,
        user_id="admin",
    )

    assert result["generation_number"] == 2
    assert orphan.read_bytes() == b"orphan-must-not-be-overwritten"
    assert {item["generation_number"] for item in result["documents"]} == {2}
    assert all("v002" in item["file_name"] for item in result["documents"])
