from __future__ import annotations

import io
import zipfile
from decimal import Decimal

import pytest
from openpyxl import Workbook

from webapp.infonalia_webapp.justificaciones_baja import imports as product_imports
from webapp.infonalia_webapp.justificaciones_baja.imports import (
    ProductImportError,
    inspect_xlsx,
    parse_decimal,
    parse_product_rows,
    preview_tabular,
    preview_xlsx,
)


MAPPING = {
    "name": "A",
    "characteristics": "B",
    "quantity": "C",
    "offered_unit_price": "D",
    "offered_amount": "E",
}


def workbook_bytes(
    rows: list[list[object]],
    *,
    sheet_name: str = "Oferta",
    other_sheet: tuple[str, list[list[object]]] | None = None,
) -> bytes:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = sheet_name
    for row in rows:
        sheet.append(row)
    if other_sheet:
        name, values = other_sheet
        extra = workbook.create_sheet(name)
        for row in values:
            extra.append(row)
    output = io.BytesIO()
    workbook.save(output)
    workbook.close()
    return output.getvalue()


def append_archive_member(content: bytes, name: str, payload: bytes = b"x") -> bytes:
    output = io.BytesIO(content)
    with zipfile.ZipFile(output, "a", compression=zipfile.ZIP_DEFLATED) as archive:
        archive.writestr(name, payload)
    return output.getvalue()


def product_rows() -> list[list[object]]:
    return [
        ["Producto", "Características", "Cantidad", "Precio", "Importe"],
        ["PAN", "80 G", 10, 0.19, 1.90],
        ["LECHE", "1 L", 5, 1.2, 6],
    ]


def test_inspect_xlsx_lists_sheets_and_returns_bounded_raw_preview() -> None:
    content = workbook_bytes(
        product_rows(),
        other_sheet=("Segundo lote", [["Producto", "Cantidad"], ["ACEITE", 2]]),
    )

    result = inspect_xlsx(content, filename="oferta.xlsx", preview_rows=2)

    assert result["format"] == "xlsx"
    assert result["sheets"] == ["Oferta", "Segundo lote"]
    assert result["sheet"] == "Oferta"
    assert result["rows"] == [
        ["Producto", "Características", "Cantidad", "Precio", "Importe"],
        ["PAN", "80 G", "10", "0.19", "1.9"],
    ]
    assert result["row_count"] == 3
    assert result["column_count"] == 5
    assert len(result["source_sha256"]) == 64
    assert "products" not in result


def test_preview_xlsx_selects_requested_sheet_and_maps_letters() -> None:
    content = workbook_bytes(
        [["irrelevante"]],
        other_sheet=("Lote 2", product_rows()),
    )

    result = preview_xlsx(
        content,
        filename="oferta.XLSX",
        sheet_name="Lote 2",
        start_row=2,
        mapping=MAPPING,
    )

    assert result["sheet"] == "Lote 2"
    assert result["mapping"] == {
        "name": 0,
        "quantity": 2,
        "offered_unit_price": 3,
        "characteristics": 1,
        "offered_amount": 4,
    }
    assert [item["name"] for item in result["products"]] == ["PAN", "LECHE"]
    assert result["can_confirm"] is True


def test_xlsx_line_ids_are_stable_for_same_file_sheet_and_rows() -> None:
    content = workbook_bytes(product_rows())

    first = preview_xlsx(content, start_row=2, mapping=MAPPING)
    second = preview_xlsx(content, start_row=2, mapping=MAPPING)

    assert [item["line_id"] for item in first["products"]] == [
        item["line_id"] for item in second["products"]
    ]
    assert first["products"][0]["line_id"].startswith("IMP-00002-")


def test_duplicate_product_names_are_allowed_but_line_ids_remain_distinct() -> None:
    content = workbook_bytes(
        [
            ["Producto", "Características", "Cantidad", "Precio", "Importe"],
            ["PAN", "45 G", 10, 0.17, 1.7],
            ["PAN", "80 G", 20, 0.19, 3.8],
        ]
    )

    result = preview_xlsx(content, start_row=2, mapping=MAPPING)

    assert [item["name"] for item in result["products"]] == ["PAN", "PAN"]
    assert len({item["line_id"] for item in result["products"]}) == 2
    assert result["can_confirm"] is True


def test_empty_and_total_rows_are_ignored_without_hiding_valid_duplicates() -> None:
    content = workbook_bytes(
        [
            ["Producto", "Características", "Cantidad", "Precio", "Importe"],
            ["PAN", "45 G", 10, 0.17, 1.7],
            [None, None, None, None, None],
            ["TOTAL GENERAL", None, None, None, 1.7],
            ["PAN", "80 G", 20, 0.19, 3.8],
            ["Subtotal lote", None, None, None, 5.5],
        ]
    )

    result = preview_xlsx(content, start_row=2, mapping=MAPPING)

    assert [item["source_row"] for item in result["products"]] == [2, 5]
    assert result["ignored_rows"] == [3, 4, 6]
    assert result["issues"] == []


def test_matching_supplied_amount_is_preserved_as_audit_input() -> None:
    result = preview_xlsx(
        workbook_bytes(product_rows()[:2]),
        start_row=2,
        mapping=MAPPING,
    )

    product = result["products"][0]
    assert product["offered_amount_input"] == "1.9"
    assert product["offered_amount_calculated"] == "1.90"
    assert product["offered_amount_origin"] == "aportado"
    assert result["issues"] == []


def test_mismatched_supplied_amount_is_a_non_blocking_warning() -> None:
    content = workbook_bytes(
        [
            ["Producto", "Características", "Cantidad", "Precio", "Importe"],
            ["PAN", "45 G", 10, 0.17, 99],
        ]
    )

    result = preview_xlsx(content, start_row=2, mapping=MAPPING)

    assert result["can_confirm"] is True
    assert result["products"][0]["offered_amount_input"] == "99"
    assert result["products"][0]["offered_amount_calculated"] == "1.70"
    assert result["issues"] == [
        {
            "row_number": 2,
            "code": "importe_no_coincide",
            "message": "El importe indicado no coincide con cantidad por precio; se conserva para revisión.",
            "severity": "advertencia",
        }
    ]


def test_invalid_optional_amount_warns_and_uses_calculated_amount() -> None:
    content = workbook_bytes(
        [
            ["Producto", "Características", "Cantidad", "Precio", "Importe"],
            ["PAN", "45 G", 10, 0.17, "importe pendiente"],
        ]
    )

    result = preview_xlsx(content, start_row=2, mapping=MAPPING)

    assert result["can_confirm"] is True
    assert result["products"][0]["offered_amount_input"] is None
    assert result["products"][0]["offered_amount_origin"] == "calculado"
    assert result["issues"][0]["code"] == "importe_invalido"
    assert result["issues"][0]["severity"] == "advertencia"


@pytest.mark.parametrize(
    ("row", "message"),
    [
        (["PAN", "45 G", True, 0.17, 1.7], "cantidad"),
        (["PAN", "45 G", 10, False, 1.7], "precio"),
        (["PAN", "45 G", 10, 0.17, True], "booleano"),
        ([True, "45 G", 10, 0.17, 1.7], "producto"),
        (["PAN", False, 10, 0.17, 1.7], "características"),
    ],
)
def test_boolean_product_values_are_rejected(row: list[object], message: str) -> None:
    content = workbook_bytes(
        [["Producto", "Características", "Cantidad", "Precio", "Importe"], row]
    )

    result = preview_xlsx(content, start_row=2, mapping=MAPPING)

    assert result["can_confirm"] is False
    assert result["products"] == []
    assert message.casefold() in result["issues"][0]["message"].casefold()


def test_required_invalid_decimal_blocks_confirmation_but_does_not_raise() -> None:
    content = workbook_bytes(
        [
            ["Producto", "Características", "Cantidad", "Precio", "Importe"],
            ["PAN", "45 G", "diez", "0,17", "1,70"],
        ]
    )

    result = preview_xlsx(content, start_row=2, mapping=MAPPING)

    assert result["products"] == []
    assert result["issues"][0]["code"] == "decimal_invalido"
    assert result["can_confirm"] is False


def test_negative_quantity_or_price_blocks_the_row() -> None:
    rows = [
        ["PAN", "", "-1", "0,17"],
        ["LECHE", "", "1", "-0,50"],
    ]

    result = parse_product_rows(
        rows,
        mapping={"name": 0, "characteristics": 1, "quantity": 2, "offered_unit_price": 3},
        start_row=1,
        source_key="test",
    )

    assert result["products"] == []
    assert [issue["code"] for issue in result["issues"]] == ["valor_negativo", "valor_negativo"]
    assert result["can_confirm"] is False


@pytest.mark.parametrize(
    ("value", "expected"),
    [
        ("1.234,56", Decimal("1234.56")),
        ("1234,56 €", Decimal("1234.56")),
        ("EUR 1.234,56", Decimal("1234.56")),
        ("1.234", Decimal("1234")),
        ("0.123", Decimal("0.123")),
        ("1234.56", Decimal("1234.56")),
        (".75", Decimal("0.75")),
        (1234, Decimal("1234")),
        (12.5, Decimal("12.5")),
        (Decimal("0.120"), Decimal("0.120")),
    ],
)
def test_parse_decimal_accepts_spanish_and_canonical_values(value: object, expected: Decimal) -> None:
    assert parse_decimal(value, field="importe") == expected


@pytest.mark.parametrize(
    "value",
    [
        None,
        True,
        "",
        "1,234.56",
        "1.2.3",
        "1e3",
        "NaN",
        "Infinity",
        "=1+1",
        object(),
    ],
)
def test_parse_decimal_rejects_ambiguous_formula_and_unsupported_values(value: object) -> None:
    with pytest.raises(ProductImportError):
        parse_decimal(value, field="importe")


@pytest.mark.parametrize("value", [Decimal("1E+1000"), Decimal("1E-1000")])
def test_parse_decimal_rejects_extreme_fixed_point_values_without_formatting_them(value: Decimal) -> None:
    with pytest.raises(ProductImportError, match="precisión admitida"):
        parse_decimal(value, field="cantidad")


def test_decimal_strings_remain_canonical_and_preserve_useful_scale() -> None:
    result = preview_tabular(
        "PAN\t10,000\t0,120\n",
        mapping={"name": "A", "quantity": "B", "offered_unit_price": "C"},
    )

    assert result["products"][0]["quantity"] == "10.000"
    assert result["products"][0]["offered_unit_price"] == "0.120"
    assert result["products"][0]["offered_amount_calculated"] == "1.200000"


def test_calculated_amount_exceeding_precision_is_a_row_error() -> None:
    result = parse_product_rows(
        [["PRODUCTO", Decimal("1E+64"), Decimal("1E+64")]],
        mapping={"name": 0, "quantity": 1, "offered_unit_price": 2},
        start_row=1,
        source_key="precision-limit",
    )

    assert result["products"] == []
    assert [issue["code"] for issue in result["issues"]] == ["decimal_invalido"]
    assert result["can_confirm"] is False


def test_preview_tabular_parses_crlf_and_preserves_stable_ids() -> None:
    text = (
        "Producto\tCaracterísticas\tCantidad\tPrecio\tImporte\r\n"
        "PAN\t45 G\t10\t0,17\t1,70\r\n"
        "PAN\t80 G\t20\t0,19\t3,80\r\n"
    )

    first = preview_tabular(text, start_row=2, mapping=MAPPING)
    second = preview_tabular(text, start_row=2, mapping=MAPPING)

    assert first["format"] == "tabular"
    assert [item["name"] for item in first["products"]] == ["PAN", "PAN"]
    assert len({item["line_id"] for item in first["products"]}) == 2
    assert [item["line_id"] for item in first["products"]] == [
        item["line_id"] for item in second["products"]
    ]


def test_tabular_formula_text_is_rejected_before_mapping() -> None:
    with pytest.raises(ProductImportError, match="fórmulas"):
        preview_tabular("PAN\t=1+1\t0,17")


def test_xlsx_formula_is_rejected_instead_of_using_a_cached_value() -> None:
    content = workbook_bytes(
        [
            ["Producto", "Características", "Cantidad", "Precio", "Importe"],
            ["PAN", "45 G", 10, 0.17, "=C2*D2"],
        ]
    )

    with pytest.raises(ProductImportError, match="fórmulas"):
        preview_xlsx(content, start_row=2, mapping=MAPPING)


def test_xlsx_error_cell_is_rejected() -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.append(["Producto", "Características", "Cantidad", "Precio", "Importe"])
    sheet.append(["PAN", "45 G", 10, 0.17, None])
    sheet["E2"] = "#DIV/0!"
    sheet["E2"].data_type = "e"
    output = io.BytesIO()
    workbook.save(output)
    workbook.close()

    with pytest.raises(ProductImportError, match="errores de Excel"):
        preview_xlsx(output.getvalue(), start_row=2, mapping=MAPPING)


@pytest.mark.parametrize("filename", ["oferta.xls", "oferta.xlsm", "oferta.csv", "../oferta.xlsx", ""])
def test_unsupported_or_unsafe_filenames_are_rejected(filename: str) -> None:
    with pytest.raises(ProductImportError):
        preview_xlsx(workbook_bytes(product_rows()), filename=filename)


def test_corrupt_non_zip_and_non_xlsx_archives_are_rejected() -> None:
    with pytest.raises(ProductImportError, match="XLSX válido"):
        preview_xlsx(b"not a zip")

    fake = io.BytesIO()
    with zipfile.ZipFile(fake, "w") as archive:
        archive.writestr("document.txt", "not an xlsx")
    with pytest.raises(ProductImportError, match="estructura mínima"):
        preview_xlsx(fake.getvalue())


@pytest.mark.parametrize(
    "member",
    [
        "../escape.xml",
        "/absolute.xml",
        "C:/escape.xml",
        "folder\\..\\escape.xml",
    ],
)
def test_xlsx_archive_rejects_unsafe_internal_paths(member: str) -> None:
    content = append_archive_member(workbook_bytes(product_rows()), member)

    with pytest.raises(ProductImportError, match="ruta interna no segura"):
        preview_xlsx(content)


@pytest.mark.parametrize(
    "member",
    [
        "xl/vbaProject.bin",
        "xl/connections.xml",
        "xl/embeddings/oleObject1.bin",
        "xl/externalLinks/externalLink1.xml",
    ],
)
def test_xlsx_archive_rejects_macro_connection_and_embedded_parts(member: str) -> None:
    content = append_archive_member(workbook_bytes(product_rows()), member)

    with pytest.raises(ProductImportError, match="no admitido"):
        preview_xlsx(content)


def test_xlsx_rejects_compressed_size_expansion_limit(monkeypatch: pytest.MonkeyPatch) -> None:
    content = workbook_bytes(product_rows())
    monkeypatch.setattr(product_imports, "MAX_ARCHIVE_UNCOMPRESSED_BYTES", 100)

    with pytest.raises(ProductImportError, match="descomprimido"):
        preview_xlsx(content)


def test_xlsx_rejects_member_count_limit(monkeypatch: pytest.MonkeyPatch) -> None:
    content = workbook_bytes(product_rows())
    monkeypatch.setattr(product_imports, "MAX_ARCHIVE_MEMBERS", 1)

    with pytest.raises(ProductImportError, match="demasiadas partes"):
        preview_xlsx(content)


def test_xlsx_rejects_upload_size_limit(monkeypatch: pytest.MonkeyPatch) -> None:
    content = workbook_bytes(product_rows())
    monkeypatch.setattr(product_imports, "MAX_IMPORT_BYTES", len(content) - 1)

    with pytest.raises(ProductImportError, match="tamaño máximo"):
        preview_xlsx(content)


def test_xlsx_rejects_row_and_column_limits(monkeypatch: pytest.MonkeyPatch) -> None:
    too_many_rows = workbook_bytes([["A"], ["B"], ["C"], ["D"]])
    monkeypatch.setattr(product_imports, "MAX_IMPORT_ROWS", 3)
    with pytest.raises(ProductImportError, match="máximo de filas"):
        preview_xlsx(too_many_rows)

    monkeypatch.setattr(product_imports, "MAX_IMPORT_ROWS", 5_000)
    monkeypatch.setattr(product_imports, "MAX_IMPORT_COLUMNS", 3)
    too_many_columns = workbook_bytes([["A", "B", "C", "D"]])
    with pytest.raises(ProductImportError, match="máximo de columnas"):
        preview_xlsx(too_many_columns)


def test_tabular_rejects_size_row_column_and_nul_limits(monkeypatch: pytest.MonkeyPatch) -> None:
    monkeypatch.setattr(product_imports, "MAX_IMPORT_BYTES", 3)
    with pytest.raises(ProductImportError, match="tamaño"):
        preview_tabular("PAN\t1")

    monkeypatch.setattr(product_imports, "MAX_IMPORT_BYTES", 10 * 1024 * 1024)
    monkeypatch.setattr(product_imports, "MAX_IMPORT_ROWS", 2)
    with pytest.raises(ProductImportError, match="máximo de filas"):
        preview_tabular("A\nB\nC")

    monkeypatch.setattr(product_imports, "MAX_IMPORT_ROWS", 5_000)
    monkeypatch.setattr(product_imports, "MAX_IMPORT_COLUMNS", 2)
    with pytest.raises(ProductImportError, match="máximo de columnas"):
        preview_tabular("A\tB\tC")

    with pytest.raises(ProductImportError, match="caracteres no permitidos"):
        preview_tabular("A\x00B")


def test_preview_is_capped_without_affecting_parse_of_all_rows() -> None:
    rows = [["Producto", "Características", "Cantidad", "Precio", "Importe"]]
    rows.extend([[f"P{index}", "", 1, 1, 1] for index in range(1, 61)])

    result = preview_xlsx(
        workbook_bytes(rows),
        start_row=2,
        mapping=MAPPING,
        preview_rows=999,
    )

    assert len(result["rows"]) == product_imports.MAX_PREVIEW_ROWS
    assert len(result["products"]) == 60


@pytest.mark.parametrize(
    "mapping",
    [
        {"name": 0, "quantity": 1},
        {"name": 0, "quantity": 1, "offered_unit_price": 1},
        {"name": 0, "quantity": 1, "offered_unit_price": 2, "unknown": 3},
        {"name": True, "quantity": 1, "offered_unit_price": 2},
        {"name": 0, "quantity": -1, "offered_unit_price": 2},
        {"name": 0, "quantity": 1, "offered_unit_price": "XFE"},
    ],
)
def test_invalid_mappings_are_rejected(mapping: dict[str, object]) -> None:
    with pytest.raises(ProductImportError):
        parse_product_rows(
            [["PAN", 1, 2]],
            mapping=mapping,
            start_row=1,
            source_key="test",
        )


@pytest.mark.parametrize("start_row", [0, -1, True, "no", 5_001])
def test_invalid_start_rows_are_rejected(start_row: object) -> None:
    with pytest.raises(ProductImportError, match="fila inicial"):
        preview_tabular("PAN\t1\t2", start_row=start_row)


@pytest.mark.parametrize("preview_rows", [0, -1, True, "no"])
def test_invalid_preview_limits_are_rejected(preview_rows: object) -> None:
    with pytest.raises(ProductImportError, match="previsualización|previsualizarse"):
        preview_tabular("PAN\t1\t2", preview_rows=preview_rows)


def test_missing_sheet_and_invalid_sheet_type_are_rejected() -> None:
    content = workbook_bytes(product_rows())

    with pytest.raises(ProductImportError, match="no existe"):
        preview_xlsx(content, sheet_name="Otro")
    with pytest.raises(ProductImportError, match="no es válida"):
        preview_xlsx(content, sheet_name=123)  # type: ignore[arg-type]


def test_import_does_not_generate_costs_or_lock_products() -> None:
    result = preview_xlsx(
        workbook_bytes(product_rows()),
        start_row=2,
        mapping=MAPPING,
    )

    for product in result["products"]:
        assert product["applied_percentage"] is None
        assert product["applied_factor"] is None
        assert product["generated_unit_cost"] is None
        assert product["manual_unit_cost"] is None
        assert product["locked"] is False
        assert product["cost_origin"] == "sin_generar"
