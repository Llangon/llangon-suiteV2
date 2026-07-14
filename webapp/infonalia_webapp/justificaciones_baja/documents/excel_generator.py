"""Pure frozen-value Excel audit generation from DocumentPayloadV1."""

from __future__ import annotations

import hashlib
from decimal import Decimal
from pathlib import Path

from openpyxl import Workbook
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from .filenames import next_versioned_path, publish_atomic_no_overwrite, temporary_output_path
from .generation import DocumentGenerationResult
from .metadata import clear_workbook_metadata
from .payload import DocumentPayloadV1
from .template_manifest import WORD_TEMPLATE_FILENAME, WORD_TEMPLATE_VERSION
from .validators import (
    DocumentValidationError,
    InvalidRouteImageError,
    inspect_route_image,
    validate_excel,
    validate_payload,
)


ACCENT = "B53613"
ACCENT_DARK = "7A2108"
ACCENT_LIGHT = "FCE7DF"
GRID = "D9B7AA"
WHITE = "FFFFFF"


def generate_excel(
    payload: DocumentPayloadV1,
    output_directory: str | Path,
    *,
    route_image_path: str | Path | None = None,
    version: int | None = None,
) -> DocumentGenerationResult:
    """Generate and atomically publish a formula-free audit workbook."""

    validate_payload(payload).require_valid()
    if payload.control.template_version != WORD_TEMPLATE_VERSION:
        raise ValueError("El payload no corresponde a la versión documental productiva.")
    image_path, warnings = _validated_image(payload, route_image_path)
    final_path, selected_version = next_versioned_path(
        output_directory,
        prefix="Auditoria_Baja",
        expediente=payload.identification.expediente,
        lot_number=payload.identification.lot_number,
        suffix=".xlsx",
        version=version,
    )
    temporary = temporary_output_path(final_path)
    try:
        workbook = _build_workbook(payload, image_path, warnings)
        workbook.save(temporary)
        workbook.close()
        report = validate_excel(temporary, payload)
        if not report.is_valid:
            raise DocumentValidationError(report)
        publish_atomic_no_overwrite(temporary, final_path)
        final_report = validate_excel(final_path, payload)
        if not final_report.is_valid:
            final_path.unlink(missing_ok=True)
            raise DocumentValidationError(final_report)
        return DocumentGenerationResult(
            path=final_path,
            sha256=_sha256_file(final_path),
            size_bytes=final_path.stat().st_size,
            warnings=tuple(warnings),
            template_version=payload.control.template_version,
            snapshot_sha256=payload.control.snapshot_sha256,
            payload_sha256=payload.sha256,
            version=selected_version,
        )
    finally:
        temporary.unlink(missing_ok=True)


def _build_workbook(
    payload: DocumentPayloadV1,
    image_path: Path | None,
    generation_warnings: list[str],
) -> Workbook:
    workbook = Workbook()
    identification = workbook.active
    identification.title = "Identificación"
    products = workbook.create_sheet("Productos")
    transport = workbook.create_sheet("Transporte")
    summary = workbook.create_sheet("Resumen")
    audit = workbook.create_sheet("Auditoría")

    _build_identification_sheet(identification, payload)
    _build_products_sheet(products, payload)
    _build_transport_sheet(transport, payload, image_path)
    _build_summary_sheet(summary, payload)
    _build_audit_sheet(audit, payload, generation_warnings)
    clear_workbook_metadata(workbook)
    workbook.calculation.calcMode = "manual"
    workbook.calculation.fullCalcOnLoad = False
    workbook.calculation.forceFullCalc = False
    return workbook


def _build_identification_sheet(sheet: object, payload: DocumentPayloadV1) -> None:
    _title(sheet, "AUDITORÍA DOCUMENTAL - IDENTIFICACIÓN", 4)
    identification = payload.identification
    rows = [
        ("Expediente", identification.expediente),
        ("Organismo", identification.organismo),
        ("Objeto", identification.objeto),
        ("Lote", identification.lot_number),
        ("Nombre del lote", identification.lot_name),
        ("Duración", identification.duration_description),
        ("Cliente", identification.client),
        ("NIF", identification.nif),
        ("Representante", identification.representative),
        ("Firmante", identification.signatory),
        ("Lugar", identification.place),
        ("Fecha", identification.date_text),
        ("Aviso", payload.narrative.estimated_draft_notice),
    ]
    for row, (label, value) in enumerate(rows, start=3):
        sheet.cell(row, 1, label)
        sheet.cell(row, 2, value)
        _style_label(sheet.cell(row, 1))
        sheet.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
        sheet.cell(row, 2).alignment = Alignment(wrap_text=True, vertical="top")
    sheet.column_dimensions["A"].width = 24
    for column in ("B", "C", "D"):
        sheet.column_dimensions[column].width = 28
    _page_setup(sheet, orientation="portrait", fit_height=1)


def _build_products_sheet(sheet: object, payload: DocumentPayloadV1) -> None:
    headers = (
        "line_id",
        "Producto",
        "Características",
        "Cantidad",
        "Precio ofertado",
        "Importe ofertado",
        "Coste generado",
        "Coste manual",
        "Coste efectivo",
        "Importe coste",
        "Margen",
        "Origen coste",
        "Bloqueado",
        "Advertencias por línea",
    )
    _title(sheet, "AUDITORÍA DOCUMENTAL - PRODUCTOS", len(headers))
    for column, header in enumerate(headers, start=1):
        sheet.cell(2, column, header)
    _style_header(sheet, 2, len(headers))
    for row, product in enumerate(payload.products, start=3):
        warnings = "; ".join(issue.code for issue in product.warnings)
        values = (
            product.line_id,
            product.name,
            product.characteristics,
            _decimal(product.quantity.raw),
            _decimal(product.offered_unit_price.raw),
            _decimal(product.offered_amount.raw),
            _decimal(product.generated_unit_cost_raw),
            _decimal(product.manual_unit_cost_raw),
            _decimal(product.effective_unit_cost.raw),
            _decimal(product.cost_amount.raw),
            _decimal(product.margin.raw),
            product.cost_origin,
            product.locked,
            warnings,
        )
        for column, value in enumerate(values, start=1):
            sheet.cell(row, column, value)
            sheet.cell(row, column).alignment = Alignment(
                horizontal="left" if column in (1, 2, 3, 12, 14) else "right",
                vertical="top",
                wrap_text=column in (2, 3, 14),
            )
        for column in (5, 7, 8, 9):
            sheet.cell(row, column).number_format = "0.000 [$€-es-ES]"
        for column in (6, 10, 11):
            sheet.cell(row, column).number_format = "#,##0.00 [$€-es-ES]"
        sheet.cell(row, 4).number_format = "#,##0.###"
    last_row = max(2, 2 + len(payload.products))
    table = Table(displayName="JustificacionProductos", ref=f"A2:N{last_row}")
    table.tableStyleInfo = TableStyleInfo(
        name="TableStyleMedium2",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    sheet.add_table(table)
    widths = (12, 26, 36, 14, 17, 18, 17, 16, 17, 18, 16, 16, 12, 32)
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width
    sheet.freeze_panes = "A3"
    sheet.auto_filter.ref = f"A2:N{last_row}"
    sheet.print_title_rows = "1:2"
    _page_setup(sheet, orientation="landscape", fit_height=0)


def _build_transport_sheet(
    sheet: object,
    payload: DocumentPayloadV1,
    image_path: Path | None,
) -> None:
    _title(sheet, "AUDITORÍA DOCUMENTAL - TRANSPORTE", 8)
    value = payload.transport
    rows = [
        ("Observatorio", value.observatory, "text"),
        ("Fecha", value.observatory_date, "text"),
        ("URL", value.observatory_url, "text"),
        ("Vehículo", value.vehicle, "text"),
        ("Semanas", value.operational_weeks, "integer"),
        ("Entregas semanales", value.weekly_deliveries, "integer"),
        ("Servicios", value.total_services, "integer"),
        ("Kilómetros ruta", _decimal(value.circular_kilometres.raw), "decimal"),
        ("Horas ruta", _decimal(value.effective_decimal_hours.raw), "decimal"),
        ("Texto humano", value.route_duration_text, "text"),
        ("Tarifa €/km", _decimal(value.kilometre_rate.raw), "rate"),
        ("Tarifa €/hora", _decimal(value.hourly_rate.raw), "rate"),
        ("Coste temporal", _decimal(value.temporal_cost.raw), "money"),
        ("Coste kilométrico", _decimal(value.kilometre_cost.raw), "money"),
        ("Coste ruta completa", _decimal(value.full_route_cost.raw), "money"),
        ("Paradas contrato", value.contract_stops, "integer"),
        ("Pedidos compartidos", value.shared_orders, "integer"),
        ("Porcentaje narrativo", _decimal(value.narrative_percentage.raw) / Decimal("100"), "percentage"),
        ("Transporte imputado", _decimal(value.allocated_transport.raw), "money"),
    ]
    for row, (label, cell_value, kind) in enumerate(rows, start=3):
        sheet.cell(row, 1, label)
        sheet.cell(row, 2, cell_value)
        _style_label(sheet.cell(row, 1))
        if kind == "integer":
            sheet.cell(row, 2).number_format = "#,##0"
        elif kind == "decimal":
            sheet.cell(row, 2).number_format = "#,##0.0000"
        elif kind == "rate":
            sheet.cell(row, 2).number_format = "0.0000 [$€-es-ES]"
        elif kind == "money":
            sheet.cell(row, 2).number_format = "#,##0.00 [$€-es-ES]"
        elif kind == "percentage":
            sheet.cell(row, 2).number_format = "0.00%"
    if image_path is not None:
        image = ExcelImage(str(image_path))
        image.width = 470
        image.height = round(470 * payload.transport.route_image.height_px / payload.transport.route_image.width_px)
        sheet.add_image(image, "E3")
    else:
        sheet["E3"] = "[IMAGEN DE RUTA PENDIENTE]"
        sheet["E3"].font = Font(name="Arial", size=10, italic=True, color=ACCENT_DARK)
    sheet.column_dimensions["A"].width = 28
    sheet.column_dimensions["B"].width = 46
    for column in ("C", "D"):
        sheet.column_dimensions[column].width = 4
    for column in ("E", "F", "G", "H"):
        sheet.column_dimensions[column].width = 18
    _page_setup(sheet, orientation="landscape", fit_height=1)


def _build_summary_sheet(sheet: object, payload: DocumentPayloadV1) -> None:
    _title(sheet, "AUDITORÍA DOCUMENTAL - RESUMEN ECONÓMICO", 4)
    headers = ("Concepto", "Valor numérico", "Valor canónico", "Presentación Word")
    for column, header in enumerate(headers, start=1):
        sheet.cell(3, column, header)
    _style_header(sheet, 3, 4)
    rows = (
        ("Oferta", payload.summary.offer, "money"),
        ("Oferta de líneas", payload.summary.justified_lines_offer, "money"),
        ("Coste crudo productos", payload.summary.raw_product_cost, "money"),
        ("Coste prorrateado productos", payload.summary.prorated_product_cost, "money"),
        ("Margen bruto", payload.summary.gross_margin, "money"),
        ("Margen bruto %", payload.summary.gross_margin_percentage, "percentage"),
        ("Transporte", payload.summary.allocated_transport, "money"),
        ("Costes indirectos", payload.summary.indirect_costs, "money"),
        ("Gastos generales", payload.summary.general_expenses, "money"),
        ("Coste total", payload.summary.total_cost, "money"),
        ("Beneficio", payload.summary.profit, "money"),
        ("Beneficio %", payload.summary.profit_percentage, "percentage"),
        ("Suma visible productos", payload.summary.visible_product_cost_sum, "money"),
        ("Residual visible", payload.summary.visual_product_residual, "money"),
    )
    thin = Side(style="thin", color=GRID)
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for row, (label, value, kind) in enumerate(rows, start=4):
        numeric = _decimal(value.raw)
        sheet.cell(row, 1, label)
        sheet.cell(row, 2, numeric)
        sheet.cell(row, 3, value.raw)
        sheet.cell(row, 4, value.display)
        sheet.cell(row, 2).number_format = "0.00%" if kind == "percentage" else "#,##0.00 [$€-es-ES]"
        for column in range(1, 5):
            cell = sheet.cell(row, column)
            cell.border = border
            cell.fill = PatternFill("solid", fgColor=ACCENT_LIGHT if row % 2 == 0 else WHITE)
    for column, width in enumerate((30, 22, 30, 24), start=1):
        sheet.column_dimensions[get_column_letter(column)].width = width
    sheet.freeze_panes = "A4"
    _page_setup(sheet, orientation="portrait", fit_height=1)


def _build_audit_sheet(
    sheet: object,
    payload: DocumentPayloadV1,
    generation_warnings: list[str],
) -> None:
    _title(sheet, "AUDITORÍA DOCUMENTAL - TRAZABILIDAD", 4)
    entries: list[tuple[str, object]] = [
        ("payload_schema_version", payload.control.payload_schema_version),
        ("snapshot_schema_version", payload.control.snapshot_schema_version),
        ("calculation_algorithm_version", payload.control.calculation_algorithm_version),
        ("template_version", payload.control.template_version),
        ("template_logical_name", WORD_TEMPLATE_FILENAME),
        ("snapshot_sha256", payload.control.snapshot_sha256),
        ("payload_sha256", payload.sha256),
        ("route_image_sha256", payload.transport.route_image.sha256 if payload.transport.route_image else ""),
        ("generated_at", payload.control.generated_at),
        ("generated_by", payload.control.generated_by),
        ("draft", payload.control.draft),
        ("product_count", len(payload.products)),
        ("generation_trace", "DocumentPayloadV1 -> XLSX; sin recálculo económico"),
        ("generation_warnings", "; ".join(generation_warnings)),
        ("economic_warnings", "; ".join(issue.code for issue in payload.warnings.economic_issues)),
        ("document_warnings", "; ".join(issue.code for issue in payload.warnings.document_warnings)),
        ("pending_validation_fields", "; ".join(payload.warnings.pending_validation_fields)),
    ]
    entries.extend(
        (f"summary.{name}", getattr(payload.summary, name).raw)
        for name in payload.summary.__dataclass_fields__
    )
    transport_raw_fields = (
        "circular_kilometres",
        "effective_decimal_hours",
        "kilometre_rate",
        "hourly_rate",
        "temporal_cost",
        "kilometre_cost",
        "full_route_cost",
        "narrative_percentage",
        "allocated_transport",
    )
    entries.extend(
        (f"transport.{name}", getattr(payload.transport, name).raw)
        for name in transport_raw_fields
    )
    for index, product in enumerate(payload.products):
        prefix = f"product.{index:04d}.{product.line_id}"
        entries.extend(
            (
                (f"{prefix}.line_id", product.line_id),
                (f"{prefix}.quantity", product.quantity.raw),
                (f"{prefix}.offered_unit_price", product.offered_unit_price.raw),
                (f"{prefix}.offered_amount", product.offered_amount.raw),
                (f"{prefix}.generated_unit_cost", product.generated_unit_cost_raw or ""),
                (f"{prefix}.manual_unit_cost", product.manual_unit_cost_raw or ""),
                (f"{prefix}.effective_unit_cost", product.effective_unit_cost.raw),
                (f"{prefix}.cost_amount", product.cost_amount.raw),
                (f"{prefix}.margin", product.margin.raw),
                (f"{prefix}.cost_origin", product.cost_origin),
                (f"{prefix}.locked", str(product.locked).lower()),
            )
        )
    for row, (key, value) in enumerate(entries, start=3):
        sheet.cell(row, 1, key)
        sheet.cell(row, 2, value)
        _style_label(sheet.cell(row, 1))
        sheet.merge_cells(start_row=row, start_column=2, end_row=row, end_column=4)
        sheet.cell(row, 2).alignment = Alignment(wrap_text=True, vertical="top")
    sheet.column_dimensions["A"].width = 38
    sheet.column_dimensions["B"].width = 42
    sheet.column_dimensions["C"].width = 24
    sheet.column_dimensions["D"].width = 24
    sheet.freeze_panes = "A3"
    _page_setup(sheet, orientation="landscape", fit_height=0)


def _title(sheet: object, title: str, columns: int) -> None:
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=columns)
    cell = sheet.cell(1, 1, title)
    cell.font = Font(name="Arial", size=16, bold=True, color=WHITE)
    cell.fill = PatternFill("solid", fgColor=ACCENT)
    cell.alignment = Alignment(horizontal="left", vertical="center")
    sheet.row_dimensions[1].height = 28
    sheet.sheet_view.showGridLines = False


def _style_header(sheet: object, row: int, columns: int) -> None:
    for column in range(1, columns + 1):
        cell = sheet.cell(row, column)
        cell.font = Font(name="Arial", size=9, bold=True, color=WHITE)
        cell.fill = PatternFill("solid", fgColor=ACCENT)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _style_label(cell: object) -> None:
    cell.font = Font(name="Arial", size=9, bold=True, color=ACCENT_DARK)


def _page_setup(sheet: object, *, orientation: str, fit_height: int) -> None:
    sheet.page_setup.orientation = orientation
    sheet.page_setup.paperSize = sheet.PAPERSIZE_A4
    sheet.sheet_properties.pageSetUpPr.fitToPage = True
    sheet.page_setup.fitToWidth = 1
    sheet.page_setup.fitToHeight = fit_height
    sheet.page_margins.left = 0.3
    sheet.page_margins.right = 0.3
    sheet.page_margins.top = 0.5
    sheet.page_margins.bottom = 0.5


def _validated_image(
    payload: DocumentPayloadV1,
    route_image_path: str | Path | None,
) -> tuple[Path | None, list[str]]:
    reference = payload.transport.route_image
    if route_image_path is None:
        if reference is not None:
            raise InvalidRouteImageError("El payload declara imagen, pero no se proporcionó el archivo.")
        return None, ["imagen_ruta_ausente"]
    if reference is None:
        raise InvalidRouteImageError("Se proporcionó una imagen no declarada en el payload.")
    path = Path(route_image_path)
    actual = inspect_route_image(path, logical_name=reference.logical_name)
    if (
        actual.logical_name,
        actual.mime_type,
        actual.width_px,
        actual.height_px,
        actual.sha256,
        actual.size_bytes,
    ) != (
        reference.logical_name,
        reference.mime_type,
        reference.width_px,
        reference.height_px,
        reference.sha256,
        reference.size_bytes,
    ):
        raise InvalidRouteImageError("La imagen no coincide con la referencia lógica del payload.")
    return path, []


def _decimal(value: str | None) -> Decimal | None:
    return Decimal(value) if value is not None else None


def _sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as stream:
        for chunk in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest().upper()


__all__ = ("generate_excel",)
